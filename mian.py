import tkinter as tk
from tkinter import ttk
from tkinter import *
from openpyxl import *



window=tk.Tk()
window.title("Data Entry")
frame=tk.Frame(window)
frame.pack()


user_info=tk.LabelFrame(frame,text="User Entry",padx=20,pady=10)
user_info.grid(row=0,column=0)
name=tk.Label(user_info,text="Enter Name")
name.grid(row=0,column=0)

name_entry=tk.Entry(user_info)
name_entry.grid(row=1,column=0)


user_search=tk.LabelFrame(frame,text="User Search",padx=20,pady=10)
user_search.grid(row=0,column=1)
sname=tk.Label(user_search,text="Enter Name")
sname.grid(row=0,column=0,sticky=(W))

sname_entry=tk.Entry(user_search)
sname_entry.grid(row=1,column=0,sticky=(W))

my_list = tk.Listbox(user_search, width=50)
my_list.grid(row=2,rowspan=2)

month=tk.Label(user_search,text="Select Month")
month.grid(row=0,column=1)

amt_label=tk.Label(user_search,text="Enter Amount")
amt_label.grid(row=2,column=1,sticky=(S,W))

amt_entry=tk.Entry(user_search)
amt_entry.grid(row=3,column=1,sticky=(N,W))


drop=ttk.Combobox(user_search,values=["January","February","March","April","May","June","July","August","September","October","November","Decamber"])
drop.grid(row=1,column=1)


window.geometry("800x400")

def addvalue():
    for cell in sheet['A']:
        if cell.value==sname_entry.get():
            x=cell.row
    for cell in sheet['1']:
        if cell.value==drop.get():
            y=cell.column
            print(y)
    sheet.cell(row=x,column=y).value=int(amt_entry.get())
    workbook.save(filename="Book1.xlsx")
    popup = tk.Toplevel(window)
    popup.title("Congratulations!")
    popup.geometry("200x200")
    pop=tk.Label(popup,text="Amount Added")
    pop.place(x=100,y=100,anchor=tk.CENTER)





def update(data):
    my_list.delete(0,END)
    for item in data:
        my_list.insert(END, item)

def fillout(e):
    sname_entry.delete(0,END)
    sname_entry.insert(0,my_list.get(ACTIVE))

def check(e):
    typed=sname_entry.get()
    if typed=="":
        data=users
    else:
        data=[]
        for item in users:
            if typed.lower() in item.lower():
                data.append(item)
    update(data)

def new_entry():
    new_line = sheet.max_row + 1
    sheet.cell(column=1, row=new_line, value=name_entry.get())
    workbook.save(filename="Book1.xlsx")


try:
    workbook = load_workbook(filename="Book1.xlsx")
    sheet = workbook.active
except:
    workbook = Workbook()
    sheet = workbook.active

sheet["A1"] = "Name"
sheet["B1"] = "January"
sheet["C1"] = "February"
sheet["D1"] = "March"
sheet["E1"] = "April"
sheet["F1"] = "May"
sheet["G1"] = "June"
sheet["H1"] = "July"
sheet["I1"] = "August"
sheet["J1"] = "September"
sheet["K1"] = "October"
sheet["L1"] = "November"
sheet["M1"] = "December"

users=[]
i=0;
for i,row in enumerate(sheet):
    if i!=0:
        users.append(row[0].value)
print(users)

update(users)

drop.bind("<<ComboboxSelected>>")

my_list.bind("<<ListboxSelect>>",fillout)
sname_entry.bind("<KeyRelease>",check)

btn=tk.Button(user_info,text="Confirm",command = new_entry)
btn.grid(row=1,column=1)

btn2=tk.Button(user_search,text="Confirm",command=addvalue)
btn2.grid(row=3,column=2,sticky=(N,W))


window.mainloop()